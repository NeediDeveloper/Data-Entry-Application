import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
import os
from tkinter import messagebox

def Save():
    try:
        Name = str(NameEntry.get())
        Marks = int(MarksEntry.get())
        Roll = int(RollEntry.get())
        Cell = int(CellEntry.get())
        
        data = {
            "Name":[Name],
            "Marks":[Marks],
            "Roll" :[Roll],
            "Cell" :[Cell]
        }
        df = pd.DataFrame(data)
        File_path = "Student_Info.xlsx"
        
        if os.path.exists(File_path):
            with pd.ExcelWriter(File_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                writer._book = load_workbook(File_path)
                
                df.to_excel(writer, startrow=writer.sheets["Sheet1"].max_row, index=False, header=False)
                writer._save()
            
            
        else:
            df.to_excel(File_path, index=False)
            
        messagebox.showinfo("Success", "Data saved successfully..!")
    except ValueError:
        messagebox.showinfo("Invalid Input", "Please Check Your Input..!")
    except Exception as e:
        messagebox.showinfo("Some Error Occurd", e)
        
        

window = tk.Tk()
window.geometry("350x400")
window.configure(bg="#282929")
window.title("Student Info")

FirstLabel = tk.Label(window, text="Student Info", bg="#282929", fg="#F3F3F3",font=("STENCIL",20))
FirstLabel.place(x=80,y=15)

NameLabel = tk.Label(window, text="Student Name :", bg="#282929", fg="#F3F3F3",font=("Britannic Bold",15))
NameLabel.place(x=10, y=80)
NameEntry = tk.Entry(window, bg="#3b3b3b", fg="#F3F3F3", font=("Britannic Bold",10))
NameEntry.place(x=160, y=83, height=25, width=150)

MarksLabel = tk.Label(window, text="Student Marks :", bg="#282929", fg="#F3F3F3",font=("Britannic Bold",15))
MarksLabel.place(x=10, y=140)
MarksEntry = tk.Entry(window, bg="#3b3b3b", fg="#F3F3F3", font=("Britannic Bold",10))
MarksEntry.place(x=160, y=143, height=25, width=150)

RollLabel = tk.Label(window, text="Student Roll :", bg="#282929", fg="#F3F3F3",font=("Britannic Bold",15))
RollLabel.place(x=10, y=200)
RollEntry = tk.Entry(window, bg="#3b3b3b", fg="#F3F3F3", font=("Britannic Bold",10))
RollEntry.place(x=160, y=203, height=25, width=150)

CellLabel = tk.Label(window, text="Student Cell :", bg="#282929", fg="#F3F3F3",font=("Britannic Bold",15))
CellLabel.place(x=10, y=260)
CellEntry = tk.Entry(window, bg="#3b3b3b", fg="#F3F3F3", font=("Britannic Bold",10))
CellEntry.place(x=160, y=263, height=25, width=150)

SubmitButton = tk.Button(window, text="Save", command=Save, activebackground="#121212",
                         activeforeground="#F3F3F3",
                         bg="#3b3b3b", fg="#F3F3F3", font=("Britannic Bold",10))
SubmitButton.place(x=100, y=320, height=25, width=130)


window.mainloop()